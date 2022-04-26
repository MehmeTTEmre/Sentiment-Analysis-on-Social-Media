import datetime
import xlsxwriter
import tweepy
import emoji
from sklearn.model_selection import train_test_split
from tensorflow.keras.preprocessing.text import Tokenizer
from tensorflow.keras.preprocessing.sequence import pad_sequences
from tensorflow.keras.layers import LSTM, Dense, Dropout, Embedding
import matplotlib.pyplot as plt
from nltk.corpus import stopwords
from keras.models import Sequential
from nltk.tokenize import RegexpTokenizer
from pandas.core.reshape.reshape import get_dummies
from sklearn.metrics import accuracy_score
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import re
import nltk
import string
import warnings
import pandas as pd
plt.style.use('ggplot')
nltk.download('punkt')
nltk.download('stopwords')
nltk.download('wordnet')
warnings.filterwarnings("ignore")

# Enter your own credentials obtained
# from your developer account
consumer_key = "your access key"
consumer_secret = "your access key"
access_key = "your access key"
access_secret = "your access key"
auth = tweepy.OAuthHandler(consumer_key, consumer_secret)
auth.set_access_token(access_key, access_secret)
api = tweepy.API(auth)

date_object = datetime.date.today()
words = "AnkaraUni"
date_since = date_object

# function to display data of each tweet
def printtweetdata(n, ith_tweet):
    print()
    print(f"Tweet {n}:")
    print(f"Username:{ith_tweet[0]}")
    print(f"Description:{ith_tweet[1]}")
    print(f"Location:{ith_tweet[2]}")
    print(f"Following Count:{ith_tweet[3]}")
    print(f"Follower Count:{ith_tweet[4]}")
    print(f"Total Tweets:{ith_tweet[5]}")
    print(f"Retweet Count:{ith_tweet[6]}")
    print(f"Tweet Text:{ith_tweet[7]}")
    print(f"Hashtags Used:{ith_tweet[8]}")

def cleantText(text):
    text = re.sub(r"@[A-Za-z0-9]+", "", text) # Removed @mentions
    text = re.sub(r"#", "", text) # Removed the "# symbol"
    text = re.sub(r"RT[\s]+", "", text) # Removed RT
    text = re.sub(r"https?:\/\/\S+", "", text) # Removed the hyper link     
    text = text.lstrip() # Removing the leading whitespace from the string.
    text = text.rstrip() # Removing the trailing whitespace from the string.
    text = text.lstrip("_") # Removing the leading underscore from the string.
    return text

# Removing all the emojis from the text.
def remove_emoji(text):
    return emoji.get_emoji_regexp().sub(u'', text)

# Creating DataFrame using pandas
db = pd.DataFrame(columns=['username', 'description', 'location', 'following', 'followers', 'totaltweets', 'retweetcount', 'text', 'hashtags', 'tweet_id'])
new_words = words + " -filter:retweets"
# We are using .Cursor() to search through twitter for the required tweets.
# The number of tweets can be restricted using .items(number of tweets)
tweets = tweepy.Cursor(api.search_tweets, q=new_words, since=date_since, tweet_mode="extended").items(100)
# .Cursor() returns an iterable object. Each item in
# the iterator has various attributes that you can access to
# get information about each tweet
list_tweets = [tweet for tweet in tweets]
    
# Counter to maintain Tweet Count
i = 1
    
array_tweets=[]
array_tweets_id=[]
array_tweets_username = []
# we will iterate over each tweet in the list for extracting information about each tweet
for tweet in list_tweets:
    username = tweet.user.screen_name
    description = tweet.user.description
    location = tweet.user.location
    following = tweet.user.friends_count
    followers = tweet.user.followers_count
    totaltweets = tweet.user.statuses_count
    retweetcount = tweet.retweet_count
    hashtags = tweet.entities['hashtags']
    tweet_id = tweet.id

    # Retweets can be distinguished by a retweeted_status attribute,
    # in case it is an invalid reference, except block will be executed
    try:
        text = tweet.retweeted_status.text
    except AttributeError:
        text = tweet.full_text
    hashtext = list()
    for j in range(0, len(hashtags)):
        hashtext.append(hashtags[j]['text'])
        
    # Here we are appending all the extracted information in the DataFrame
    ith_tweet = [username, description, location, following, followers, totaltweets, retweetcount, text, hashtext, tweet_id]
    db.loc[len(db)] = ith_tweet

    if (not tweet.retweeted) and ('RT @' not in tweet.full_text):
        array_tweets.append(remove_emoji(cleantText(ith_tweet[7])))
        array_tweets_id.append(ith_tweet[9])
        array_tweets_username.append(ith_tweet[0])
        
path = "data/Daily_Tweets/" + date_object.strftime("%A") + ".xlsx"
# Create a workbook and add a worksheet.
workbook = xlsxwriter.Workbook(path)
worksheet = workbook.add_worksheet()

# Add a bold format to use to highlight cells.
bold = workbook.add_format({'bold': 1})

# Adjust the column width.
worksheet.set_column(0, 1, 20) 
worksheet.set_column(2, 2, 280) 
worksheet.set_column(3, 3, 10) 
	
# Write some data headers.
worksheet.write('A1', 'ID', bold)
worksheet.write('B1', 'Username', bold)
worksheet.write('C1', 'Tweet', bold)
worksheet.write('D1', 'Sentiment', bold)

# Start from the first cell below the headers.
row = 1
col = 0

for i in array_tweets_id:
    worksheet.write_string(row, col,str(i))
    row += 1
row = 1
for i in array_tweets_username:
    worksheet.write_string(row, col+1,str(i))
    row += 1
row = 1
for i in array_tweets:
    worksheet.write_string(row, col+2,str(i))
    row += 1
workbook.close()

df = pd.read_excel('data/Tweets_train2.xlsx')
df = df[["Sentiment", "Text"]]

# Changing the value of the column "Sentiment" to 0 if the value is "Negatif".
df["Sentiment"][df["Sentiment"]== "Negatif"] = 0
df["Sentiment"][df["Sentiment"]== "Pozitif"] = 1

# Filtering the dataframe by the value of the column "Sentiment" and it is getting the rows that have the value of 1.
df_positive = df[df["Sentiment"] == 1]
df_negative = df[df["Sentiment"] == 0]  

# Getting the first 1000 rows of the dataframe.
df_negative = df_negative.iloc[:int(2000)]
df_positive = df_positive.iloc[:int(2000)]

# Concatenating the dataframes.
df = pd.concat([df_positive, df_negative])
df["Text"] = df["Text"].str.lower() 

df2 = pd.read_excel(path)
df2 = df2[["Tweet"]]

# Cleaning and removing stopwords from list
stop_words = list(stopwords.words('turkish'))
", ".join(stopwords.words('turkish'))
STOPWORDS = set(stopwords.words('turkish'))
def cleaning_stopwords(text):
    return " ".join([word for word in str(text).split() if word not in STOPWORDS])
df['Text'] = df["Text"].apply(lambda text: cleaning_stopwords(text))
df2["Tweet"] = df2["Tweet"].apply(lambda text: cleaning_stopwords(text))

# Cleaning and removing punctuations
turkish_punctuations = string.punctuation
punctuations_list = turkish_punctuations
def cleaning_punctutations(text):
    translator = str.maketrans("", "", punctuations_list)
    return text.translate(translator)
df["Text"] = df["Text"].apply(lambda x: cleaning_punctutations(x))
df2["Tweet"] = df2["Tweet"].apply(lambda x: cleaning_punctutations(x))

# Cleaning and removing repeating characters
def cleaning_repeating_char(text):
    return re.sub(r"(.)\1+", r"\1", text)
df["Text"] = df["Text"].apply(lambda x: cleaning_repeating_char(x))
df2["Tweet"] = df2["Tweet"].apply(lambda x: cleaning_repeating_char(x))

# Cleaning and removing numeric numbers
def cleaning_numbers(data):
    return re.sub("[0-9]", "", str(data))
df["Text"] = df["Text"].apply(lambda x: cleaning_numbers(x))
df2["Tweet"] = df2["Tweet"].apply(lambda x: cleaning_numbers(x))

wb = load_workbook(filename=path)
ws = wb.worksheets[0]
for i in range(len(df2.Tweet)):
    ws["C"+str(i+2)] = df2.Tweet[i]
wb.save(path)
wb.close()

# Getting Tokenization of tweet text
tokenizer = RegexpTokenizer(r"\w+")
df["Text"] = df["Text"].apply(tokenizer.tokenize)
df2["Tweet"] = df2["Tweet"].apply(tokenizer.tokenize)

# Applying stemming
st = nltk.PorterStemmer()
def stemming_on_text(data):
    text = [st.stem(word) for word in data]
    return data
df["Text"] = df["Text"].apply(lambda x: stemming_on_text(x))
df2["Tweet"] = df2["Tweet"].apply(lambda x: stemming_on_text(x))

# Applying lemmatizer
lm = nltk.WordNetLemmatizer()
def lemmatizer_on_text(data):
    text = [lm.lemmatize(word) for word in data]
    return data
df["Text"] = df["Text"].apply(lambda x: lemmatizer_on_text(x))
df2["Tweet"] = df2["Tweet"].apply(lambda x: lemmatizer_on_text(x))

X_train_verb, X_test_verb, y_train_verb, y_test_verb = train_test_split(df.Text, df.Sentiment, test_size=0.2, random_state=0)
tokenizer = Tokenizer(num_words=5000, split=",")
tokenizer.fit_on_texts(df.Text.values)
sequences = tokenizer.texts_to_sequences(df.Text.values)
X = pad_sequences(sequences, padding="post")

# Getting the 3000 most frequent words in the text.
tokenizer = Tokenizer(num_words=5000, split=",")
tokenizer.fit_on_texts(df2.Tweet.values)
sequences = tokenizer.texts_to_sequences(df2.Tweet.values)
X2 = pad_sequences(sequences, padding="post", maxlen=X.shape[1])

model = Sequential()
model.add(Embedding(5000, 256, input_length=X.shape[1]))
model.add(Dropout(0.3))
model.add(LSTM(256, return_sequences=True, dropout=0.3, recurrent_dropout=0.2))
model.add(LSTM(256, dropout=0.3, recurrent_dropout=0.2))
model.add(Dense(2, activation="softmax"))
model.compile(loss="categorical_crossentropy", optimizer="adam", metrics=["accuracy"])
model.summary()

y = get_dummies(df.Sentiment).values
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=0)
model.fit(X_train, y_train, epochs=8, batch_size=32, verbose=2)
y_pred = model.predict(X_test)
y_pred = (y_pred > 0.5)

accr1 = accuracy_score(y_test, y_pred)

predictions = model.predict(X2)
predictions = (predictions > 0.5)
pred = []
for i in range(len(predictions)):
    if predictions[0][0] == True:
        pred.append("Negative")
    else:
        pred.append("Pozitive")

wb = load_workbook(filename=path)
ws = wb.worksheets[0]
for i in range(len(predictions)):
  ws["D"+str(i+2)] = pred[i]
wb.save(path)
wb.close()