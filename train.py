from sklearn.model_selection import train_test_split
from tensorflow.keras.preprocessing.text import Tokenizer
from tensorflow.keras.preprocessing.sequence import pad_sequences
from tensorflow.keras.layers import LSTM, Dense, Dropout, Embedding
import matplotlib.pyplot as plt
from nltk.corpus import stopwords
from keras.models import Sequential
from nltk.tokenize import RegexpTokenizer
from pandas.core.reshape.reshape import get_dummies
import numpy as np
from sklearn.metrics import accuracy_score
from PIL import Image
import matplotlib.pyplot as plt
from sklearn.metrics import ConfusionMatrixDisplay
from openpyxl import load_workbook
import sqlite3
from itertools import cycle
from sklearn.metrics import roc_curve, auc
import re
from sklearn.metrics import confusion_matrix
import nltk
import string
import warnings
import pandas as pd
plt.style.use('ggplot')
nltk.download('punkt')
nltk.download('stopwords')
nltk.download('wordnet')
warnings.filterwarnings("ignore")

df = pd.read_excel('data/train/Tweets_train.xlsx')
df.rename(columns={"Tweet" : "Text", "Etiket" : "Sentiment"}, inplace = True)
df = df[["Text", "Sentiment"]]

df2 = pd.read_excel('data/TwitterSentimentAnalysis.xlsx')
df2 = df2[["Tweet"]]

# Changing the value of the column "Sentiment" to 1 if the value is "kızgın".
df["Sentiment"][df["Sentiment"]== "kızgın"] = 1
df["Sentiment"][df["Sentiment"]== "korku"] = 2
df["Sentiment"][df["Sentiment"]== "mutlu"] = 3
df["Sentiment"][df["Sentiment"]== "surpriz"] = 4
df["Sentiment"][df["Sentiment"]== "üzgün"] = 5

# Filtering the dataframe by the value of the column "Sentiment" and it is getting the rows that have the value of 1.
df_angry = df[df["Sentiment"] == 1]
df_fear = df[df["Sentiment"] == 2]   
df_happy = df[df["Sentiment"] == 3] 
df_shocked = df[df["Sentiment"] == 4]  
df_sad = df[df["Sentiment"] == 5] 

# Getting the first 100 rows of the dataframe.
df_angry = df_angry.iloc[:int(100)]
df_fear = df_fear.iloc[:int(100)]
df_happy = df_happy.iloc[:int(100)]
df_shocked = df_shocked.iloc[:int(100)]
df_sad = df_sad.iloc[:int(100)]

# Concatenating the dataframes.
df = pd.concat([df_angry, df_fear, df_happy, df_shocked, df_sad])
df["Text"] = df["Text"].str.lower() 

# Cleaning and removing stopwords from list
stop_words = list(stopwords.words('turkish'))
", ".join(stopwords.words('turkish'))
STOPWORDS = set(stopwords.words('turkish'))
def cleaning_stopwords(text):
    return " ".join([word for word in str(text).split() if word not in STOPWORDS])
df['Text'] = df["Text"].apply(lambda text: cleaning_stopwords(text))
df2["Tweet"] = df2["Tweet"].apply(lambda text: cleaning_stopwords(text))

# Cleaning and removing punctuations
turkish_punctuations = string.punctuation
punctuations_list = turkish_punctuations
def cleaning_punctutations(text):
    translator = str.maketrans("", "", punctuations_list)
    return text.translate(translator)
df["Text"] = df["Text"].apply(lambda x: cleaning_punctutations(x))
df2["Tweet"] = df2["Tweet"].apply(lambda x: cleaning_punctutations(x))

# Cleaning and removing repeating characters
def cleaning_repeating_char(text):
    return re.sub(r"(.)\1+", r"\1", text)
df["Text"] = df["Text"].apply(lambda x: cleaning_repeating_char(x))
df2["Tweet"] = df2["Tweet"].apply(lambda x: cleaning_repeating_char(x))

# Cleaning and removing numeric numbers
def cleaning_numbers(data):
    return re.sub("[0-9]", "", str(data))
df["Text"] = df["Text"].apply(lambda x: cleaning_numbers(x))
df2["Tweet"] = df2["Tweet"].apply(lambda x: cleaning_numbers(x))

wb = load_workbook(filename="data/TwitterSentimentAnalysis.xlsx")
ws = wb.worksheets[0]
for i in range(len(df2.Tweet)):
    ws["C"+str(i+2)] = df2.Tweet[i]
wb.save("data/TwitterSentimentAnalysis.xlsx")
wb.close()

# Getting Tokenization of tweet text
tokenizer = RegexpTokenizer(r"\w+")
df["Text"] = df["Text"].apply(tokenizer.tokenize)
df2["Tweet"] = df2["Tweet"].apply(tokenizer.tokenize)

# Applying stemming
st = nltk.PorterStemmer()
def stemming_on_text(data):
    text = [st.stem(word) for word in data]
    return data
df["Text"] = df["Text"].apply(lambda x: stemming_on_text(x))
df2["Tweet"] = df2["Tweet"].apply(lambda x: stemming_on_text(x))

# Applying lemmatizer
lm = nltk.WordNetLemmatizer()
def lemmatizer_on_text(data):
    text = [lm.lemmatize(word) for word in data]
    return data
df["Text"] = df["Text"].apply(lambda x: lemmatizer_on_text(x))
df2["Tweet"] = df2["Tweet"].apply(lambda x: lemmatizer_on_text(x))

X_train_verb, X_test_verb, y_train_verb, y_test_verb = train_test_split(df.Text, df.Sentiment, test_size=0.2, random_state=0)
tokenizer = Tokenizer(num_words=3000, split=",")
tokenizer.fit_on_texts(df.Text.values)
sequences = tokenizer.texts_to_sequences(df.Text.values)
X = pad_sequences(sequences, padding="post")

# Getting the 3000 most frequent words in the text.
tokenizer = Tokenizer(num_words=3000, split=",")
tokenizer.fit_on_texts(df2.Tweet.values)
sequences = tokenizer.texts_to_sequences(df2.Tweet.values)
X2 = pad_sequences(sequences, padding="post", maxlen=X.shape[1])

model = Sequential()
model.add(Embedding(3000, 256, input_length=X.shape[1]))
model.add(Dropout(0.3))
model.add(LSTM(256, return_sequences=True, dropout=0.3, recurrent_dropout=0.2))
model.add(LSTM(256, dropout=0.3, recurrent_dropout=0.2))
model.add(Dense(5, activation="softmax"))
model.compile(loss="categorical_crossentropy", optimizer="adam", metrics=["accuracy"])
model.summary()

y = get_dummies(df.Sentiment).values
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=0)
model.fit(X_train, y_train, epochs=10, batch_size=32, verbose=2)
y_pred = model.predict(X_test)
y_pred = (y_pred > 0.5)
#[print(X_test_verb.values[i], y_pred[i], y_test[i]) for i in range(len(y_test))]
#accr1 = model.evaluate(X_train,y_train) #we are starting to test the model here
accr1 = accuracy_score(y_test, y_pred)

predictions = model.predict(X2)
predictions = (predictions > 0.5)
[print(df2["Tweet"].values[i], predictions[i]) for i in range(len(predictions))]
pred = []
for i in range(len(predictions)):
  if predictions[i][0] == True:
    pred.append("Sinirli")
  elif predictions[i][1] == True:
    pred.append("Korku")
  elif predictions[i][2] == True:
    pred.append("Mutlu")
  elif predictions[i][3] == True:
    pred.append("Sürpriz")
  else:
    pred.append("Üzgün")

wb = load_workbook(filename="data/TwitterSentimentAnalysis.xlsx")
ws = wb.worksheets[0]
for i in range(len(predictions)):
  ws["D"+str(i+2)] = pred[i]
wb.save("data/TwitterSentimentAnalysis.xlsx")
wb.close()

db_ex = pd.read_excel('data/TwitterSentimentAnalysis.xlsx')
db_ex = db_ex[["ID", "Username", "Tweet", "Sentiment"]]

con = sqlite3.connect("data/data.db")   
cursor = con.cursor()
cursor.execute("CREATE TABLE IF NOT EXISTS Twitter (ID TEXT, Username TEXT, Tweet TEXT, Sentiment TEXT)")
con.commit()
for i in range(len(db_ex.ID)):
    cursor.execute("INSERT INTO Twitter (ID,Username,Tweet, Sentiment) VALUES (?,?,?,?)", (str(db_ex.ID[i]),db_ex.Username[i],db_ex.Tweet[i],db_ex.Sentiment[i]))
    con.commit()
con.close()

real = []
for i in range(len(y_test)):
  real.append(np.argmax(y_test[i]))

pred = []
for i in range(len(y_pred)):
  pred.append(np.argmax(y_pred[i]))

# Compute Confusion Matrix
CR = confusion_matrix(real, pred)
labels = ["Sinirli", "Korku", "Mutlu", "Sürpriz", "Üzgün"]
disp = ConfusionMatrixDisplay(confusion_matrix=CR, display_labels=labels)
disp = disp.plot(cmap="Blues")
plt.tick_params(axis=u'both', which=u'both',length=0)
plt.grid(b=None)
plt.savefig("image/analysis/ConfusionMatrix5.jpg")
image = Image.open("image/analysis/ConfusionMatrix5.jpg")
new_image = image.resize((500, 471))
new_image.save("image/analysis/ConfusionMatrix5.jpg")
plt.clf()

# Compute ROC curve and ROC area for each class
fpr = dict()
tpr = dict()
roc_auc = dict()
for i in range(y_test.shape[1]):
    fpr[i], tpr[i], _ = roc_curve(y_test[:, i], y_pred[:, i])
    roc_auc[i] = auc(fpr[i], tpr[i])


# Compute micro-average ROC curve and ROC area
fpr["micro"], tpr["micro"], _ = roc_curve(y_test.ravel(), y_pred.ravel())
roc_auc["micro"] = auc(fpr["micro"], tpr["micro"])
colors = cycle(["red", "blue", "green","yellow","orange"])
for i, color in zip(range(y_test.shape[1]), colors):
    if i == 0:
      j = "sinirli"
    elif i == 1:
      j = "korku"
    elif i == 2:
      j = "mutlu"
    elif i == 3:
      j = "surpriz"
    else:
      j = "üzgün"
    plt.plot(
        fpr[i],
        tpr[i],
        color=color,
        label="ROC curve of class " + j + "(area = {1:0.2f})".format(i,roc_auc[i]),
    )

plt.plot([0, 1], [0, 1], "k--")
plt.xlim([0.0, 1.0])
plt.ylim([0.0, 1.05])
plt.xlabel("False Positive Rate")
plt.ylabel("True Positive Rate")
plt.title("Receiver operating characteristic to multiclass")
plt.legend(loc="lower right")
plt.savefig("image/analysis/ROC5.jpg")
image = Image.open("image/analysis/ROC5.jpg")
new_image = image.resize((500, 471))
new_image.save("image/analysis/ROC5.jpg")
plt.clf()
